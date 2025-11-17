import openpyxl
import zipfile
import os
import re
from pathlib import Path
import xml.etree.ElementTree as ET
from collections import defaultdict

def sanitize_filename(filename):
    """Remove invalid characters from filename"""
    # Remove invalid characters for Windows filenames
    invalid_chars = r'[<>:"/\\|?*]'
    filename = re.sub(invalid_chars, '_', filename)
    # Remove leading/trailing spaces and dots
    filename = filename.strip(' .')
    # Limit length to avoid issues
    if len(filename) > 200:
        filename = filename[:200]
    return filename

def get_file_extension_from_bytes(data):
    """Determine file extension from image bytes"""
    if len(data) >= 8 and data[:8] == b'\x89PNG\r\n\x1a\n':
        return '.png'
    elif len(data) >= 2 and data[:2] == b'\xff\xd8':
        return '.jpg'
    elif len(data) >= 6 and data[:6] in (b'GIF87a', b'GIF89a'):
        return '.gif'
    elif len(data) >= 2 and data[:2] == b'BM':
        return '.bmp'
    elif len(data) >= 4 and data[:4] in (b'II*\x00', b'MM\x00*'):
        return '.tiff'
    return '.png'  # default

def extract_images_from_excel(excel_file, output_dir='extracted_images'):
    """
    Extract images from column C and rename them with product names from column B.
    Based on the approach that XLSX files are ZIP archives with images in xl/media/
    """
    # Create output directory if it doesn't exist
    Path(output_dir).mkdir(exist_ok=True)
    
    # Load the workbook to read cell data
    print(f"Loading Excel file: {excel_file}")
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    print(f"Working with sheet: {ws.title}")
    
    # Dictionary to store images mapped by row number
    images_by_row = {}
    
    # Open Excel file as ZIP archive (XLSX files are ZIP archives)
    print("\nExtracting images from Excel archive structure...")
    try:
        with zipfile.ZipFile(excel_file, 'r') as zip_ref:
            # Get all image files from xl/media/ directory
            image_files = sorted([f for f in zip_ref.namelist() if f.startswith('xl/media/')])
            print(f"Found {len(image_files)} image file(s) in xl/media/ folder")
            
            if not image_files:
                print("No images found in the Excel file.")
                return
            
            # Find all sheet XML files
            sheet_files = sorted([f for f in zip_ref.namelist() 
                                 if f.startswith('xl/worksheets/sheet') and f.endswith('.xml')])
            
            # Process each sheet to find image positions
            for sheet_idx, sheet_path in enumerate(sheet_files):
                sheet_num = sheet_idx + 1
                print(f"\nProcessing sheet {sheet_num}: {os.path.basename(sheet_path)}")
                
                # Read sheet XML to find drawing references
                sheet_xml = zip_ref.read(sheet_path)
                sheet_root = ET.fromstring(sheet_xml)
                
                # Namespaces for Excel XML
                ns = {
                    'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                }
                
                # Find drawing relationship
                drawing_elem = sheet_root.find('.//main:drawing', ns)
                if drawing_elem is None:
                    continue
                
                drawing_rId = drawing_elem.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                if not drawing_rId:
                    continue
                
                # Read sheet relationships to find drawing file
                sheet_rels_path = f'xl/worksheets/_rels/{os.path.basename(sheet_path)}.rels'
                if sheet_rels_path not in zip_ref.namelist():
                    continue
                
                sheet_rels_xml = zip_ref.read(sheet_rels_path)
                sheet_rels_root = ET.fromstring(sheet_rels_xml)
                rels_ns = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                
                drawing_path = None
                for rel in sheet_rels_root.findall('.//r:Relationship', rels_ns):
                    if rel.get('Id') == drawing_rId:
                        drawing_path = f"xl/{rel.get('Target')}"
                        break
                
                if not drawing_path or drawing_path not in zip_ref.namelist():
                    continue
                
                # Read drawing XML to get image positions
                drawing_xml = zip_ref.read(drawing_path)
                draw_root = ET.fromstring(drawing_xml)
                
                # Namespaces for drawing XML
                draw_ns = {
                    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                }
                
                # Read drawing relationships to map embed IDs to image files
                drawing_rels_path = f'xl/drawings/_rels/{os.path.basename(drawing_path)}.rels'
                embed_to_image = {}
                
                if drawing_rels_path in zip_ref.namelist():
                    drawing_rels_xml = zip_ref.read(drawing_rels_path)
                    drawing_rels_root = ET.fromstring(drawing_rels_xml)
                    
                    for rel in drawing_rels_root.findall('.//r:Relationship', rels_ns):
                        rel_id = rel.get('Id')
                        rel_target = rel.get('Target')
                        if rel_target:
                            # Convert relative path to absolute
                            if not rel_target.startswith('media/'):
                                image_path = f"xl/{rel_target}"
                            else:
                                image_path = f"xl/{rel_target}"
                            embed_to_image[rel_id] = image_path
                
                # Find all image anchors and their positions
                for anchor in draw_root.findall('.//xdr:twoCellAnchor', draw_ns):
                    from_elem = anchor.find('.//xdr:from', draw_ns)
                    if from_elem is None:
                        continue
                    
                    col_elem = from_elem.find('.//xdr:col', draw_ns)
                    row_elem = from_elem.find('.//xdr:row', draw_ns)
                    
                    if col_elem is None or row_elem is None:
                        continue
                    
                    col = int(col_elem.text) + 1  # Convert 0-based to 1-based
                    row = int(row_elem.text) + 1
                    
                    # Only process images in column C (column 3)
                    if col != 3:
                        continue
                    
                    # Find the image reference
                    blip = anchor.find('.//a:blip', draw_ns)
                    if blip is None:
                        continue
                    
                    embed_id = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                    if not embed_id:
                        continue
                    
                    # Get the image file path
                    image_path = embed_to_image.get(embed_id)
                    if not image_path or image_path not in zip_ref.namelist():
                        continue
                    
                    # Read the image data
                    image_data = zip_ref.read(image_path)
                    images_by_row[row] = image_data
                    print(f"  Mapped image to row {row}, column C: {os.path.basename(image_path)}")
            
            # Fallback: If no images were mapped via XML, try matching by order
            if not images_by_row and image_files:
                print("\nNo images mapped via XML structure. Attempting fallback matching by row order...")
                # Get all rows with product names in column B
                rows_with_products = []
                for row in range(1, ws.max_row + 1):
                    cell_b = ws.cell(row=row, column=2)
                    if cell_b.value:
                        rows_with_products.append(row)
                
                # Match images to rows in order
                for idx, image_path in enumerate(image_files):
                    if idx < len(rows_with_products):
                        row = rows_with_products[idx]
                        images_by_row[row] = zip_ref.read(image_path)
                        print(f"  Matched image {idx+1} to row {row} (by order): {os.path.basename(image_path)}")
    
    except Exception as e:
        print(f"Error accessing Excel archive: {str(e)}")
        import traceback
        traceback.print_exc()
        return
    
    # Extract and rename images
    print(f"\n{'='*60}")
    print(f"Extracting {len(images_by_row)} image(s)...")
    print(f"{'='*60}")
    
    extracted_count = 0
    skipped_count = 0
    
    for row_num, image_data in sorted(images_by_row.items()):
        try:
            # Get product name from column B
            product_name_cell = ws.cell(row=row_num, column=2)
            product_name = str(product_name_cell.value) if product_name_cell.value else f"product_row_{row_num}"
            
            # Sanitize the product name
            sanitized_name = sanitize_filename(product_name)
            
            if not sanitized_name or sanitized_name == "None":
                sanitized_name = f"product_row_{row_num}"
            
            # Determine file extension
            extension = get_file_extension_from_bytes(image_data)
            
            # Create filename
            filename = f"{sanitized_name}{extension}"
            filepath = os.path.join(output_dir, filename)
            
            # Handle duplicate filenames
            counter = 1
            while os.path.exists(filepath):
                base_name = sanitized_name
                filename = f"{base_name}_{counter}{extension}"
                filepath = os.path.join(output_dir, filename)
                counter += 1
            
            # Save the image
            with open(filepath, 'wb') as f:
                f.write(image_data)
            
            print(f"[OK] Extracted: {filename} (row {row_num})")
            extracted_count += 1
            
        except Exception as e:
            print(f"[ERROR] Error processing row {row_num}: {str(e)}")
            import traceback
            traceback.print_exc()
            skipped_count += 1
    
    print(f"\n{'='*60}")
    print(f"Extraction complete!")
    print(f"Successfully extracted: {extracted_count} images")
    print(f"Skipped: {skipped_count} images")
    print(f"Output directory: {os.path.abspath(output_dir)}")
    print(f"{'='*60}")

if __name__ == "__main__":
    excel_file = "Hermizon images.xlsx"
    extract_images_from_excel(excel_file)
